from io import BytesIO, StringIO
import pandas as pd
import requests
import json
import os
import io
from pathlib import Path
import zipfile
import openpyxl as xls
from bs4 import BeautifulSoup
from datetime import datetime
from typing import Dict, List, Tuple, Optional

SEP: str = ";"
DECIMAL: str = ","


def get_download_link() -> str:
    base_url = "https://energy.ec.europa.eu"
    response = requests.get(f"{base_url}/data-and-analysis/weekly-oil-bulletin_en")
    response.raise_for_status()
    soup = BeautifulSoup(response.text, "html.parser")

    ecl_file_divs = soup.find_all("div", class_="ecl-file")
    assert ecl_file_divs

    ecl_file_div = None

    for div in ecl_file_divs:
        container = div.find("div", class_="ecl-file__container")
        if not container:
            continue

        info = container.find("div", class_="ecl-file__info")
        if not info:
            continue

        if "onwards" in info.text:
            ecl_file_div = div
            break

    assert ecl_file_div

    ecl_file_action_div = ecl_file_div.find("div", class_="ecl-file__action")
    assert ecl_file_action_div

    file_download_a = ecl_file_action_div.find(
        "a",
        class_="ecl-link ecl-link--standalone ecl-link--icon ecl-file__download",
    )
    assert file_download_a

    donwload_link = f"{base_url}{file_download_a.get('href')}"

    return donwload_link


def download(download_link: str) -> xls.Workbook:
    response = requests.get(download_link)
    response.raise_for_status()
    xlsx_file = BytesIO(response.content)
    workbook = xls.load_workbook(xlsx_file)
    return workbook


def extract_dates(workbook: xls.Workbook) -> List[datetime]:
    worksheet = workbook["Prices with taxes"]
    date_column = 1

    dates = []
    current_row = 4
    current_date = worksheet.cell(row=current_row, column=date_column).value
    while current_date:
        dates.append(current_date)
        current_row += 1
        current_date = worksheet.cell(row=current_row, column=date_column).value

    return dates


def extract_countries(
    workbook: xls.Workbook, country_translations: Dict[str, str]
) -> Tuple[List[str], List[int]]:
    worksheet = workbook["Prices with taxes"]

    countries_de = []
    country_column_offsets = []
    current_col = 16
    country_row_offset = 4
    current_country_code = str(
        worksheet.cell(row=country_row_offset, column=current_col).value
    )
    while current_country_code != "None":
        countries_de.append(country_translations[current_country_code])
        country_column_offsets.append(current_col)

        current_col += 1
        value = str(worksheet.cell(row=country_row_offset, column=current_col).value)
        while value not in country_translations and len(value) and current_col < 2000:
            current_col += 1
            value = str(
                worksheet.cell(row=country_row_offset, column=current_col).value
            )

        current_country_code = str(
            worksheet.cell(row=country_row_offset, column=current_col).value
        )

    return countries_de, country_column_offsets


def extract_data_for_country(
    workbook: xls.Workbook,
    country_idx: int,
    date_amount: int,
    country_column_offsets: List[int],
    diesel: bool,
) -> List[float]:
    worksheet = workbook["Prices with taxes"]
    data_row_offset = 4
    data_factor = 0.001

    country_column_offset = country_column_offsets[country_idx]
    header_1 = str(worksheet.cell(row=1, column=country_column_offset + 1).value)
    if "exchange" in header_1:
        data = [
            worksheet.cell(
                row=data_row_offset + i,
                column=country_column_offset + (3 if diesel else 2),
            ).value
            for i in range(date_amount)
        ]
    else:
        data = [
            worksheet.cell(
                row=data_row_offset + i,
                column=country_column_offset + (2 if diesel else 1),
            ).value
            for i in range(date_amount)
        ]

    data = [round(data_factor * (d or float("nan")), 2) for d in data]  # type: ignore

    return data


def extract_data(
    workbook: xls.Workbook, since: Optional[datetime]
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    dates = extract_dates(workbook)
    date_counter = 0
    if since is not None:
        for date in dates:
            if date >= since:
                date_counter += 1
            else:
                break
    else:
        date_counter = len(dates)
    dates = dates[:date_counter]

    with open("country_names_code_de.json", "r", encoding="utf-8") as f:
        country_codes = json.load(f)

    countries_de, country_column_offsets = extract_countries(workbook, country_codes)

    super_df, diesel_df = None, None

    for diesel in (False, True):
        df_data = {"Tag": [date.strftime("%Y/%m/%d") for date in dates]} | {
            country_de: extract_data_for_country(
                workbook, country_idx, len(dates), country_column_offsets, diesel
            )
            for country_idx, country_de in enumerate(countries_de)
        }

        if diesel:
            diesel_df = pd.DataFrame(df_data)
        else:
            super_df = pd.DataFrame(df_data)

    assert super_df is not None
    assert diesel_df is not None

    return super_df, diesel_df


def prepare_data(
    super_df: pd.DataFrame, diesel_df: pd.DataFrame
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    with open("de_neighbor_countries.json", "r", encoding="utf-8") as f:
        neighbors = json.load(f)

    all_super_df = super_df[["Tag"] + neighbors]
    all_diesel_df = diesel_df[["Tag"] + neighbors]

    sorted_super_cols = ["Tag"] + all_super_df.iloc[0, 1:].sort_values(
        ascending=False
    ).index.tolist()
    all_super_df = all_super_df[sorted_super_cols]
    all_super_df = all_super_df.iloc[::-1].reset_index(drop=True)

    sorted_diesel_cols = ["Tag"] + all_diesel_df.iloc[0, 1:].sort_values(
        ascending=False
    ).index.tolist()
    all_diesel_df = all_diesel_df[sorted_diesel_cols]
    all_diesel_df = all_diesel_df.iloc[::-1].reset_index(drop=True)

    super_df_ = super_df.iloc[0:2]
    diesel_df_ = diesel_df.iloc[0:2]

    results = []

    value_colun_name = ""
    for df in (super_df_, diesel_df_):
        melted = df.melt(id_vars="Tag", var_name="country", value_name="value")
        year, month, day = [int(x) for x in str(melted["Tag"].iloc[0]).split("/")]
        pivoted = melted.pivot(
            index="country", columns="Tag", values="value"
        ).reset_index()
        pivoted.columns = ["country", "old", "new"]
        pivoted["diff"] = ((pivoted["new"] - pivoted["old"]) / pivoted["old"]) * 100
        pivoted["diff"] = pivoted["diff"].round(2)
        sorted_df = pivoted.sort_values(by="new", ascending=False)
        value_colun_name = (f"Euro je Liter, Stand {day:02d}.{month:02d}.{year}",)
        results.append(
            sorted_df[["country", "new", "diff"]].rename(
                columns={
                    "country": "EU-Staat",
                    "new": value_colun_name,
                    "diff": "Prozentuale Veränderung",
                }
            )
        )

    recent_super_df, recent_diesel_df = results

    super_de_row = recent_super_df[recent_super_df["EU-Staat"] == "Deutschland"]
    super_other_rows = recent_super_df[
        recent_super_df["EU-Staat"] != "Deutschland"
    ].sort_values(value_colun_name, ascending=False)
    recent_super_df = pd.concat([super_de_row, super_other_rows])
    recent_super_df = recent_super_df.reset_index(drop=True)

    diesel_de_row = recent_diesel_df[recent_diesel_df["EU-Staat"] == "Deutschland"]
    diesel_other_rows = recent_diesel_df[
        recent_diesel_df["EU-Staat"] != "Deutschland"
    ].sort_values(value_colun_name, ascending=False)
    recent_diesel_df = pd.concat([diesel_de_row, diesel_other_rows])
    recent_diesel_df = recent_diesel_df.reset_index(drop=True)

    assert isinstance(recent_super_df, pd.DataFrame)
    assert isinstance(all_super_df, pd.DataFrame)
    assert isinstance(recent_diesel_df, pd.DataFrame)
    assert isinstance(all_diesel_df, pd.DataFrame)

    return recent_super_df, all_super_df, recent_diesel_df, all_diesel_df


def generate_filename(
    all_: bool, diesel: bool, since: Optional[datetime] = None
) -> str:
    now = datetime.now()
    prefix = f"{now.day:02d}-{now.month:02d}-{now.year}_{now.hour:02d}{now.minute:02d}{now.second:02d}"
    filename = os.path.join(
        "output",
        f"{prefix}{'_{type_}_'.format(type_='diesel' if diesel else 'super')}"
        + (
            "recent"
            if not all_
            else (
                "all"
                if not since
                else f"since_{since.strftime('%Y-%m-%d')}"  # type: ignore
            )
        )
        + ".csv",
    )
    return filename


def main() -> io.BytesIO:
    download_link = get_download_link()
    workbook = download(download_link)
    since = datetime(year=2016, month=7, day=20)
    super_df, diesel_df = extract_data(workbook, since)
    recent_super_df, all_super_df, recent_diesel_df, all_diesel_df = prepare_data(
        super_df, diesel_df
    )

    filenames = []

    for df, all_, diesel in zip(
        [recent_super_df, all_super_df, recent_diesel_df, all_diesel_df],
        [False, True, False, True],
        [False, False, True, True],
    ):
        csv_io = StringIO()
        df.to_csv(csv_io, index=False, sep=SEP, decimal=DECIMAL)
        header_, *rows_ = csv_io.getvalue().split("\n")

        header = SEP.join(f'"{col}"' for col in header_.split(SEP)) + "\n"
        rows = []
        for row_ in rows_[:-1]:
            cells = row_.split(SEP)
            row = f'"{cells[0]}"{SEP}' + SEP.join(cell for cell in cells[1:]) + "\n"
            rows.append(row)

        filename = generate_filename(all_, diesel, since)
        filenames.append(filename)

        with open(filename, "w", encoding="utf-8") as f:
            f.writelines([header] + rows)

    zip_buffer = io.BytesIO()

    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
        for file_path in filenames:
            path = Path(file_path)
            zip_file.write(path, arcname=path.name)

    zip_buffer.seek(0)

    return zip_buffer


if __name__ == "__main__":
    main()
