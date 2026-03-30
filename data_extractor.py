import os
import json
import datetime as dt
import openpyxl as xls
import pandas as pd
from typing import Dict, List, Tuple, Optional


class RecentDataExtractor:

    SHEET_NAME: str = "Sheet1"
    DATE_SPLITTER: str = "/"
    DATE_CELL: Dict[str, int] = {"row": 2, "column": 1}
    COUNTRY_CELLS: List[Dict[str, int]] = [
        {"row": r, "column": 1} for r in range(3, 29 + 1)
    ]
    DATA_CELLS: List[Dict[str, int]] = [
        {"row": r, "column": 2} for r in range(3, 29 + 1)
    ]
    DATA_FACTOR: float = 0.001
    HEADER: List[str] = ["EU-Staat", "Euro je Liter, Stand {day}.{month}.{year}"]
    TRANSLATION_FILENAME: str = "country_names_en_de.json"

    def _get_country_translations(self) -> Dict[str, str]:
        with open(self.TRANSLATION_FILENAME, "r", encoding="utf-8") as f:
            return json.load(f)

    def _sort_data(
        self, countries: List[str], data: List[float]
    ) -> Tuple[List[str], List[float]]:
        zipped = list(zip(countries, data))
        sorted_zipped = sorted(zipped, key=lambda x: x[1], reverse=True)
        sorted_countries, sorted_data = zip(*sorted_zipped)
        return sorted_countries, sorted_data  # type: ignore

    def extract(self, workbook: xls.Workbook, last_df: pd.DataFrame) -> pd.DataFrame:
        worksheet = workbook[self.SHEET_NAME]

        date = worksheet.cell(**self.DATE_CELL).value
        assert isinstance(date, dt.datetime)
        header = [
            h.format(day=f"{date.day:02d}", month=f"{date.month:02d}", year=date.year)
            for h in self.HEADER
        ]

        countries_en = [
            str(worksheet.cell(**cell).value) for cell in self.COUNTRY_CELLS
        ]
        data = [
            float(f"{(float(worksheet.cell(**cell).value) * self.DATA_FACTOR):.2f}")  # type: ignore
            for cell in self.DATA_CELLS
        ]

        country_name_translations = self._get_country_translations()

        countries_de = [country_name_translations[c] for c in countries_en]

        countries_de, data = self._sort_data(countries_de, data)

        df_data = {header[0]: countries_de, header[1]: data}

        df = pd.DataFrame(df_data)

        value_column_now = [
            col for col in df.columns if col.startswith("Euro je Liter")
        ][0]
        value_column_last = [
            col for col in last_df.columns if col.startswith("Euro je Liter")
        ][0]

        merged = df.merge(last_df, on="EU-Staat", how="left")
        merged["Prozentuale Veränderung"] = round(
            (
                (merged[value_column_now] - merged[value_column_last])
                / merged[value_column_last]
            )
            * 100,
            2,
        )
        df = df.merge(
            merged[["EU-Staat", "Prozentuale Veränderung"]], on="EU-Staat", how="left"
        )

        return df


class AllDataExtractor:

    SHEET_NAME: str = "Prices with taxes"

    # dates
    DATE_COLUMN: int = 1
    DATE_ROW_OFFSET: int = 4

    # counries
    COUNTRY_ROW_OFFSET: int = 4
    COUNTRY_COLUMN_OFFSET: int = 16

    # data
    DATA_ROW_OFFSET: int = 4
    DATA_COLUMN_OFFSET: int = 17
    DATA_ROW_STEP: int = 7

    TRANSLATION_FILENAME: str = "country_names_code_de.json"

    DATA_FACTOR: float = 0.001

    def _extract_dates(self, workbook: xls.Workbook) -> List[dt.datetime]:
        worksheet = workbook[self.SHEET_NAME]

        dates = []
        current_row = self.DATE_ROW_OFFSET
        current_date = worksheet.cell(row=current_row, column=self.DATE_COLUMN).value
        while current_date:
            dates.append(current_date)
            current_row += 1
            current_date = worksheet.cell(
                row=current_row, column=self.DATE_COLUMN
            ).value

        return dates

    def _extract_countries(
        self, workbook: xls.Workbook, country_translations: Dict[str, str]
    ) -> Tuple[List[str], List[int]]:
        worksheet = workbook[self.SHEET_NAME]

        countries_de = []
        country_column_offsets = []
        current_col = self.COUNTRY_COLUMN_OFFSET
        current_country_code = str(
            worksheet.cell(row=self.COUNTRY_ROW_OFFSET, column=current_col).value
        )
        while current_country_code != "None":
            countries_de.append(country_translations[current_country_code])
            country_column_offsets.append(current_col)

            current_col += 1
            value = str(
                worksheet.cell(row=self.COUNTRY_ROW_OFFSET, column=current_col).value
            )
            while (
                value not in country_translations and len(value) and current_col < 2000
            ):
                current_col += 1
                value = str(
                    worksheet.cell(
                        row=self.COUNTRY_ROW_OFFSET, column=current_col
                    ).value
                )

            current_country_code = str(
                worksheet.cell(row=self.COUNTRY_ROW_OFFSET, column=current_col).value
            )

        return countries_de, country_column_offsets

    def _get_country_translations(self) -> Dict[str, str]:
        with open(self.TRANSLATION_FILENAME, "r", encoding="utf-8") as f:
            return json.load(f)

    def _extract_data_for_country(
        self,
        workbook: xls.Workbook,
        country_index: int,
        length: int,
        country_column_offsets: List[int],
    ) -> List[float]:
        worksheet = workbook[self.SHEET_NAME]

        country_column_offset = country_column_offsets[country_index]
        header_1 = str(worksheet.cell(row=1, column=country_column_offset + 1).value)
        if "exchange" in header_1:
            data = [
                worksheet.cell(
                    row=self.DATA_ROW_OFFSET + i, column=country_column_offset + 2
                ).value
                for i in range(length)
            ]
        else:
            data = [
                worksheet.cell(
                    row=self.DATA_ROW_OFFSET + i, column=country_column_offset + 1
                ).value
                for i in range(length)
            ]

        data = [round(self.DATA_FACTOR * (d or float("nan")), 2) for d in data]  # type: ignore

        return data

    def extract(
        self, workbook: xls.Workbook, since: Optional[dt.datetime]
    ) -> pd.DataFrame:
        dates = self._extract_dates(workbook)
        date_counter = 0
        if since is not None:
            for date in dates:
                if date >= since:
                    date_counter += 1
                else:
                    break
        else:
            date_counter = len(dates)
        dates = dates[:date_counter]

        country_translations = self._get_country_translations()
        countries_de, country_column_offsets = self._extract_countries(
            workbook, country_translations
        )

        df_data = {"Tag": [date.strftime("%Y/%m/%d") for date in dates]} | {
            country_de: self._extract_data_for_country(
                workbook, i, len(dates), country_column_offsets
            )
            for i, country_de in enumerate(countries_de)
        }

        df = pd.DataFrame(df_data)

        return df
